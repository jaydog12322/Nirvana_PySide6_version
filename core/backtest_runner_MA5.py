# backtest_runner_MA5.py
import sys
import argparse

import pandas as pd
import numpy as np
from datetime import timedelta
import os
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart import LineChart, Reference

from openpyxl.utils import get_column_letter

# Read trigger parameters from environment variables, with defaults
TRIGGER_RATE = float(os.environ.get("TRIGGER_RATE", 10))
TRIGGER_VOLUME = int(os.environ.get("TRIGGER_VOLUME", 10_000_000_000))
TRIGGER_FOREIGN = int(os.environ.get("TRIGGER_FOREIGN", 0))
TRIGGER_INSTITUTION = int(os.environ.get("TRIGGER_INSTITUTION", 0))

# -----------------------------
# Config Section
# -----------------------------

STRATEGY_ID = 'ma5_support_v1.20.7'

# Placeholder input path — fill this in manually
INPUT_FILE_PATH = 'PATH/TO/Enhanced_000000_ma5support_v1.parquet'


# -----------------------------
# Helper Functions
# -----------------------------

def calculate_target_price(entry_price, ma5, ma10):
    """Calculate target price using MA5 and MA10 spread."""
    spread_pct = (ma5 - ma10) / ma10
    return round(entry_price * (1 + spread_pct), 2)


def calculate_return(entry_price, exit_price):
    """Return % return between entry and exit prices."""
    return round((exit_price - entry_price) / entry_price * 100, 2)


def calculate_slope(series, window=3):
    """Simple slope: linear fit over window days ending at current day."""
    if len(series) < window:
        return np.nan
    x = np.arange(window)
    y = np.array(series)[-window:]

    coeffs = np.polyfit(x, y, 1)
    return round(coeffs[0], 4)  # slope


def is_valid_entry(row):
    """Returns True if entry conditions are met. Currently checks for 정배열 only."""
    return row['MA_5'] > row['MA_10'] > row['MA_20']


# -----------------------------
# Main Simulation Logic
# -----------------------------

def run_backtest(df: pd.DataFrame, stock_code: str, stock_name: str,
                 min_rate: float, max_rate: float,
                 min_volume: float, max_volume: float,
                 min_foreign: float, min_institution: float,
                 manual_trigger_date=None) -> list:
    # Convert string arguments (from GUI) into proper numeric types
    try:
        min_rate = float(min_rate) if min_rate != "" else None
        max_rate = float(max_rate) if max_rate != "" else None
        min_volume = int(min_volume) if min_volume != "" else None
        max_volume = int(max_volume) if max_volume != "" else None
        min_foreign = int(min_foreign) if min_foreign != "" else None
        min_institution = int(min_institution) if min_institution != "" else None
    except Exception as e:
        print(f"[!] Parameter type conversion error: {e}")
        return []

    results = []

    df = df.sort_values('일자').reset_index(drop=True)

    waiting_for_ma20_dip = False
    last_exit_idx = -1

    for idx, row in df.iterrows():
        if idx == 0:
            continue

        # If using manual trigger mode, skip all rows except the trigger date
        if manual_trigger_date is not None:
            if row['일자'].date() != manual_trigger_date.date():
                continue

        if waiting_for_ma20_dip:
            if row['종가'] < row['MA_20']:
                waiting_for_ma20_dip = False
            else:
                continue

        if (
                (min_rate is None or row['등락률'] > min_rate) and
                (max_rate is None or row['등락률'] < max_rate) and
                (min_volume is None or row['거래대금'] > min_volume) and
                (max_volume is None or row['거래대금'] < max_volume) and
                (min_foreign is None or row.get('외국인_순매수', 0) > min_foreign) and
                (min_institution is None or row.get('기관_순매수', 0) > min_institution) and
                df.loc[idx - 1, '등락률'] > -10 and
                row['종가'] >= row['MA_5'] * 1.05
        ):
            # 🚫 Reject trigger if MA20 is more than 4% above MA10
            ma10 = row['MA_10']
            ma20 = row['MA_20']
            if (ma20 - ma10) / ma10 * 100 > 4:
                continue

            # 🚫 Reject trigger if MA10 is more than 5% above MA5
            ma5 = row['MA_5']
            if (ma10 - ma5) / ma5 * 100 > 5:
                continue

            # 🚫 Skip if previous day's open was a severe gap-down (>10%)
            prev_row = df.loc[idx - 1]
            gap_down_pct = (row['시가'] - prev_row['종가']) / prev_row['종가'] * 100
            if gap_down_pct < -10:
                continue

            # ✅ New flat-day filter: no 5+ consecutive flat days in past 20
            past_20 = df.loc[max(0, idx - 20): idx - 1, '등락률']
            consecutive_flats = 0
            has_3_or_more_consecutive_zeros = False

            for r in past_20:
                if r == 0:
                    consecutive_flats += 1
                    if consecutive_flats >= 3:
                        has_3_or_more_consecutive_zeros = True
                        break
                else:
                    consecutive_flats = 0

            if has_3_or_more_consecutive_zeros:
                continue

            trigger_date = row['일자']
            trigger_row = row

            # === Trigger intra-day max % until MA5 touched ===
            trigger_prev_close = df.loc[idx - 1, '종가']
            trigger_max_high = row['고가']
            for i in range(idx + 1, len(df)):
                row_i = df.loc[i]
                trigger_max_high = max(trigger_max_high, row_i['고가'])
                if row_i['저가'] <= row_i['MA_5']:
                    break
            trigger_intra_high_pct = round((trigger_max_high - trigger_prev_close) / trigger_prev_close * 100, 2)

            # === Entry Logic (Updated MA5 Cross-Down Rule) ===
            entry_idx = None
            entry_price = None

            for forward_idx in range(idx + 1, len(df)):
                row_forward = df.loc[forward_idx]
                ma5 = row_forward['MA_5']
                ma10 = row_forward['MA_10']
                ma20 = row_forward['MA_20']
                low = row_forward['저가']
                high = row_forward['고가']
                open_price = row_forward['시가']
                close_price = row_forward['종가']

                # 🚩 Condition 1: any of the four prices dips below MA5
                prices = [low, high, open_price, close_price]
                crossed_down_ma5 = any(p < ma5 for p in prices)

                if crossed_down_ma5:
                    if open_price < ma10:
                        entry_price = None
                        break

                    # ✅ Improved logic: separate gap-down vs overlap cases
                    if open_price < ma5:
                        # Gap-down under MA5, but above MA10 → enter at open
                        entry_price = open_price
                    elif low <= ma5 <= high:
                        # Ideal overlap → enter at MA5
                        entry_price = ma5
                    else:
                        # Fallback → enter at open
                        entry_price = open_price

                    # 🚩 Optional filters
                    spread = (ma5 - ma10) / ma10 * 100
                    if spread < 3:
                        entry_price = None
                        break

                    if not is_valid_entry(row_forward):
                        entry_price = None
                        break

                    entry_idx = forward_idx
                    break

            if entry_idx is None or entry_price is None:
                continue

            entry_row = df.loc[entry_idx]
            entry_date = entry_row['일자']
            # entry_price = entry_row['MA_5']
            # ✅ Calculate trading-day gap between trigger and entry (excluding trigger day itself)
            days_took_to_entry = df.loc[(df['일자'] > trigger_date) & (df['일자'] <= entry_date)].shape[0]

            # ✅ Calculate pre-entry peak return (%)
            trigger_idx = idx  # The index where the trigger was found
            peak_high = df.loc[trigger_idx:entry_idx, '고가'].max()
            pre_entry_peak_return_pct = round((peak_high - entry_price) / entry_price * 100, 2)

            if not is_valid_entry(entry_row):
                continue

            # 🆕 NEW: Extract False_Entry_Checker value from entry row
            false_entry_checker = entry_row.get('False_Entry_Checker', None)

            # === Spreads & 정배열 체크 ===
            정배열_5_10_20_60 = (
                    entry_row['MA_5'] > entry_row['MA_10'] > entry_row['MA_20'] > entry_row['MA_60']
            )
            정배열_5_10_20_60_120 = (
                    entry_row['MA_5'] > entry_row['MA_10'] > entry_row['MA_20'] > entry_row['MA_60'] > entry_row[
                'MA_120']
            )
            spread_ma5_10 = round((entry_row['MA_5'] - entry_row['MA_10']) / entry_row['MA_10'] * 100, 2)
            spread_ma5_20 = round((entry_row['MA_5'] - entry_row['MA_20']) / entry_row['MA_20'] * 100, 2)

            # === Exit Logic ===
            ma10_entry = entry_row['MA_10']
            target_price = calculate_target_price(entry_price, entry_row['MA_5'], ma10_entry)
            exit_price = None
            exit_date = None
            outcome = None
            note = ""
            max_drawdown = 0

            for exit_idx in range(entry_idx, len(df)):
                row_exit = df.loc[exit_idx]
                high = row_exit['고가']
                low = row_exit['저가']
                ma10 = row_exit['MA_10']

                open_price = row_exit['시가']
                opened_below_ma10 = open_price < ma10

                drawdown = (low - entry_price) / entry_price * 100
                max_drawdown = min(max_drawdown, drawdown)

                hit_target = high >= target_price
                hit_ma10 = low <= ma10 or opened_below_ma10

                # 🚫 NEW RULE: skip same-day target-only exits (only allow same-day loss)
                if hit_target and hit_ma10:
                    # Conservative exit: MA10 hit wins over target
                    if opened_below_ma10:
                        exit_price = open_price
                        note = "both hit — gap-down below MA10 → exited at open (loss)"
                    else:
                        exit_price = ma10
                        note = "both hit — MA10 breached intraday → exited at MA10 (loss)"
                    exit_date = row_exit['일자']
                    outcome = 'loss'
                    break

                # ✅ Insert this check here to disallow same-day profit-only exits
                if exit_idx == entry_idx and hit_target and not hit_ma10:
                    continue  # Skip — don't count same-day wins

                elif hit_target:
                    exit_price = target_price
                    exit_date = row_exit['일자']
                    outcome = 'win'
                    note = "target hit first"
                    break
                elif hit_ma10:
                    if opened_below_ma10:
                        exit_price = open_price
                        note = "gap-down below MA10 → exited at open"
                    else:
                        exit_price = ma10
                        note = "intraday MA10 breach → exited at MA10"
                    exit_date = row_exit['일자']
                    outcome = 'loss'
                    break

            if exit_price is None:
                continue

            # === Max gain until MA10 touch (excluding entry day) ===
            max_high_until_ma10 = None

            # ❗ Fix: skip gain calculation if trade exited same day
            if exit_date != entry_date:
                for i in range(entry_idx + 1, len(df)):
                    row_i = df.loc[i]
                    open_price = row_i['시가']
                    low_price = row_i['저가']
                    ma10 = row_i['MA_10']

                    # 🚫 Stop before processing this row if MA10 is touched intraday
                    if low_price <= ma10:
                        break

                    # ✅ Only include day's high if opened at/above MA10
                    if open_price >= ma10:
                        high_price = row_i['고가']
                        if max_high_until_ma10 is None:
                            max_high_until_ma10 = high_price
                        else:
                            max_high_until_ma10 = max(max_high_until_ma10, high_price)

            if max_high_until_ma10 is not None:
                custom_max_gain_pct = round((max_high_until_ma10 - entry_price) / entry_price * 100, 2)
            else:
                custom_max_gain_pct = 0.0

            days_held = df[(df['일자'] > entry_date) & (df['일자'] <= exit_date)].shape[0]
            # Check for suspected trading halts
            halt_suspect_rows = df.loc[entry_idx + 1: exit_idx]
            flat_rows = halt_suspect_rows[
                (halt_suspect_rows['시가'] == halt_suspect_rows['종가']) &
                (halt_suspect_rows['시가'] == halt_suspect_rows['고가']) &
                (halt_suspect_rows['시가'] == halt_suspect_rows['저가'])
                ]

            if flat_rows.shape[0] > 1:
                note += " | 거래정지의심"

            slope_ma5 = calculate_slope(df.loc[:entry_idx]['MA_5'].values)
            slope_ma10 = calculate_slope(df.loc[:entry_idx]['MA_10'].values)
            slope_ma20 = calculate_slope(df.loc[:entry_idx]['MA_20'].values)

            return_pct = calculate_return(entry_price, exit_price)
            risk_to_reward = round(return_pct / spread_ma5_10, 4) if spread_ma5_10 else None

            results.append({
                '종목코드': stock_code,
                '종목명': stock_name,
                '테마명': df.get('manual_theme', [None])[0] if 'manual_theme' in df.columns else None,
                'trigger_date': trigger_date,
                'entry_date': entry_date,
                'exit_date': exit_date,
                'entry_price': entry_price,
                'exit_price': exit_price,
                'return_pct': return_pct,
                'days_held': days_held,
                'target_price': target_price,
                'outcome': outcome,
                'note': note,
                'max_gain_pct': custom_max_gain_pct,
                'max_drawdown': round(max_drawdown, 2),
                'spread_ma5_10': spread_ma5_10,
                'spread_ma5_20': spread_ma5_20,
                '정배열_5_10_20_60': 정배열_5_10_20_60,
                '정배열_5_10_20_60_120': 정배열_5_10_20_60_120,
                'trigger_intra_high_pct': trigger_intra_high_pct,
                'slope_ma5': slope_ma5,
                'slope_ma10': slope_ma10,
                'slope_ma20': slope_ma20,
                '외국인_순매수': trigger_row.get('외국인_순매수', None),
                '기관_순매수': trigger_row.get('기관_순매수', None),
                'entry_id': f"{stock_code}_EN{len(results) + 1}",
                'exit_id': f"{stock_code}_EX{len(results) + 1}",
                'trigger_rate_used': min_rate,
                'trigger_volume_used': min_volume,
                'trigger_foreign_used': min_foreign,
                'trigger_institution_used': min_institution,
                'risk_to_reward': risk_to_reward,
                'pre_entry_peak_return_pct': pre_entry_peak_return_pct,
                'Days_took_to_Entry': days_took_to_entry,
                'False_Entry_Checker': false_entry_checker,  # 🆕 NEW COLUMN

            })

            waiting_for_ma20_dip = True
            last_exit_idx = exit_idx

    return results


def generate_daily_summary(df_results: pd.DataFrame) -> pd.DataFrame:
    df = df_results.copy()
    df['entry_date'] = pd.to_datetime(df['entry_date'])
    df['exit_date'] = pd.to_datetime(df['exit_date'])

    # --- Simulate daily stock holdings ---
    all_dates = pd.date_range(df['entry_date'].min(), df['exit_date'].max())
    held_counts = pd.Series(0, index=all_dates)

    for _, row in df.iterrows():
        held_range = pd.date_range(row['entry_date'], row['exit_date'])
        held_counts.loc[held_range] += 1

    max_held_stocks = held_counts.max()

    # --- Group by entry date ---
    grouped = df.groupby('entry_date')
    summary = pd.DataFrame({
        'return_pct (sum)': grouped['return_pct'].sum(),
        '# of Trades (sum)': grouped.size(),
        'max_gain_pct (average)': grouped['max_gain_pct'].mean(),
        'max_drawdown (average)': grouped['max_drawdown'].mean(),
        'spread_ma5_10 (average)': grouped['spread_ma5_10'].mean(),
        'risk_to_reward (sum)': grouped['risk_to_reward'].sum(),
        'risk_to_reward (average)': grouped['risk_to_reward'].mean(),
        '# of stocks bought (sum)': grouped.size(),
    })

    # --- Count sells on each exit date ---
    exit_counts = df['exit_date'].value_counts().rename('# of stocks sold (sum)')
    summary['# of stocks sold (sum)'] = summary.index.map(exit_counts).fillna(0).astype(int)

    # --- Add overlapping held count ---
    summary['# of stocks held (sum)'] = summary.index.map(held_counts).fillna(0).astype(int)

    summary = summary.round(2)

    # --- Add Total and Average rows at the bottom ---
    total_row = summary.sum(numeric_only=True)
    total_row.name = 'Total'

    avg_row = summary.mean(numeric_only=True)
    avg_row.name = 'Average'

    summary = pd.concat([summary, pd.DataFrame([total_row, avg_row])])

    return summary


def generate_monthly_summary(df_results: pd.DataFrame) -> pd.DataFrame:
    df = df_results.copy()
    df['entry_date'] = pd.to_datetime(df['entry_date'])
    df['month'] = df['entry_date'].dt.to_period('M').astype(str)

    grouped = df.groupby('month')

    # Base metrics
    return_sum = grouped['return_pct'].sum()
    risk_avg = grouped['risk_to_reward'].mean()
    trade_counts = grouped.size()

    # Win/Loss counts
    win_counts = grouped.apply(lambda g: (g['return_pct'] > 0).sum())
    lose_counts = grouped.apply(lambda g: (g['return_pct'] <= 0).sum())
    win_ratios = (win_counts / trade_counts).round(2)

    # Build table
    monthly = pd.DataFrame({
        'return_pct (sum)': return_sum,
        'cumulative_return_pct': return_sum,
        'Win (#)': win_counts,
        'Lose (#)': lose_counts,
        'Win Ratio (Win/Total # Trade)': win_ratios,
        'risk_to_reward (average)': risk_avg,
        '# stocks traded (sum)': trade_counts
    })

    # Add cumulative return after building the DataFrame
    monthly['cumulative_return_pct'] = monthly['return_pct (sum)'].cumsum()

    monthly = monthly.round(2)
    monthly = monthly.T  # transpose

    # Add Total and Average columns
    monthly['Total'] = monthly.sum(axis=1)
    monthly['Average'] = monthly.mean(axis=1)

    return monthly


def generate_weekly_summary(df_results: pd.DataFrame) -> pd.DataFrame:
    df = df_results.copy()
    df['entry_date'] = pd.to_datetime(df['entry_date'])
    df['week'] = df['entry_date'].dt.strftime('%G-W%V')  # e.g., '2024-W27'

    grouped = df.groupby('week')

    # Compute each metric
    return_sum = grouped['return_pct'].sum()
    win_counts = grouped.apply(lambda g: (g['return_pct'] > 0).sum())
    lose_counts = grouped.apply(lambda g: (g['return_pct'] <= 0).sum())
    total_counts = grouped.size()
    win_ratios = (win_counts / total_counts).round(2)

    # Build final DataFrame
    summary = pd.DataFrame({
        'return_pct (sum)': return_sum,
        'Win (#)': win_counts,
        'Lose (#)': lose_counts,
        'Win Ratio (Win#/# of Trades)': win_ratios,
    })

    summary = summary.round(2)
    return summary


def calculate_max_streaks(df: pd.DataFrame) -> tuple:
    df = df.copy()
    df['entry_date'] = pd.to_datetime(df['entry_date'])
    daily_returns = df.groupby('entry_date')['return_pct'].sum().sort_index()

    win_streak = max_streak = 0
    max_win = 0
    for val in daily_returns:
        if val > 0:
            win_streak += 1
            max_win = max(max_win, win_streak)
        else:
            win_streak = 0

    loss_streak = max_loss = 0
    for val in daily_returns:
        if val < 0:
            loss_streak += 1
            max_loss = max(max_loss, loss_streak)
        else:
            loss_streak = 0

    return max_win, max_loss


# -----------------------------
# Batch Runner
# -----------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("parquet_path", help="Path to enhanced .parquet file")
    parser.add_argument("strategy_id", help="Strategy ID for output folder naming")
    parser.add_argument("min_rate", nargs="?", default=TRIGGER_RATE)
    parser.add_argument("max_rate", nargs="?", default=None)
    parser.add_argument("min_volume", nargs="?", default=TRIGGER_VOLUME)
    parser.add_argument("max_volume", nargs="?", default=None)
    parser.add_argument("min_foreign", nargs="?", default=TRIGGER_FOREIGN)
    parser.add_argument("min_institution", nargs="?", default=TRIGGER_INSTITUTION)
    parser.add_argument("--manual-trigger", help="Path to Excel file containing manually defined triggers")

    args = parser.parse_args()

    input_path = args.parquet_path
    strategy_id = args.strategy_id

    # ✅ Create output folder using strategy_id
    OUTPUT_DIR = f'Test_Results/{strategy_id}'
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"▶ Loading data from: {input_path}")
    if args.manual_trigger:
        print(f"▶ Manual trigger mode activated: {args.manual_trigger}")
    else:
        print(
            f"▶ Parameters: 등락률 ≥ {args.min_rate}, ≤ {args.max_rate}, 거래대금 ≥ {args.min_volume}, ≤ {args.max_volume}, 외국인 ≥ {args.min_foreign}, 기관 ≥ {args.min_institution}")

    df_all = pd.read_parquet(input_path)

    # Check required columns
    required_cols = ['일자', '종목코드', '종목명', '고가', '저가', '종가',
                     '거래대금', '등락률', 'MA_5', 'MA_10', 'MA_20']
    for col in required_cols:
        if col not in df_all.columns:
            raise ValueError(f"Missing required column: {col}")

    all_results = []

    if args.manual_trigger:

        manual_df = pd.read_excel(args.manual_trigger)
        manual_df['날짜'] = pd.to_datetime(manual_df['날짜'])
        df_all['일자'] = pd.to_datetime(df_all['일자'])

        for _, row in manual_df.iterrows():
            stock_name = row['종목명']
            trigger_date = row['날짜']
            theme = row.get('테마명', None)

            # Filter stock data
            stock_df = df_all[df_all['종목명'] == stock_name].copy()
            if stock_df.empty:
                print(f"⚠️ No data found for: {stock_name}")
                continue

            code = stock_df['종목코드'].iloc[0]
            stock_df = stock_df.sort_values('일자').reset_index(drop=True)

            # Inject theme column
            stock_df['manual_theme'] = theme
            results = run_backtest(
                stock_df, code, stock_name,
                args.min_rate, args.max_rate,
                args.min_volume, args.max_volume,
                args.min_foreign, args.min_institution,
                manual_trigger_date=trigger_date
            )
            all_results.extend(results)

    else:
        grouped = df_all.groupby(['종목코드', '종목명'])

        for (code, name), group_df in grouped:
            print(f"▶ Running backtest for {code} ({name})")
            results = run_backtest(group_df.copy(), code, name,
                                   args.min_rate, args.max_rate,
                                   args.min_volume, args.max_volume,
                                   args.min_foreign, args.min_institution)
            all_results.extend(results)

    if not all_results:
        print("⚠️ No trades triggered.")
        return

    # Remove duplicates if using manual trigger file
    if args.manual_trigger:
        before = len(all_results)
        df_temp = pd.DataFrame(all_results)
        df_temp.drop_duplicates(subset=["종목코드", "trigger_date"], inplace=True)
        after = len(df_temp)
        print(f"[Deduplication] Removed {before - after} duplicate entries.")
        all_results = df_temp.to_dict(orient="records")

    df_results = pd.DataFrame(all_results)

    # 🆕 NEW: Calculate Verified_Return as the very last step
    def calculate_verified_return(row):
        false_entry_checker = row.get('False_Entry_Checker')
        return_pct = row.get('return_pct', 0)

        if false_entry_checker == "Entry_Made":
            return return_pct
        else:  # "No_Entry_Made", None, or any other value
            return 0

    df_results['Verified_Return'] = df_results.apply(calculate_verified_return, axis=1)

    # Recalculate held_counts here for use in both summary and global stats
    all_dates = pd.date_range(df_results['entry_date'].min(), df_results['exit_date'].max())
    held_counts = pd.Series(0, index=all_dates)

    for _, row in df_results.iterrows():
        held_range = pd.date_range(row['entry_date'], row['exit_date'])
        held_counts.loc[held_range] += 1

    output_path = os.path.join(OUTPUT_DIR, f'{strategy_id}_results.xlsx')

    # Generate summary tables
    daily_summary_df = generate_daily_summary(df_results)
    monthly_summary_df = generate_monthly_summary(df_results)

    weekly_summary_df = generate_weekly_summary(df_results)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_results.to_excel(writer, index=False, sheet_name='Results')
        daily_summary_df.to_excel(writer, sheet_name='Summary', startrow=0)
        monthly_summary_df.to_excel(writer, sheet_name='Summary 2', startrow=0)
        weekly_summary_df.to_excel(writer, sheet_name='Summary 3', startrow=0)

        # --- Weekly Win Ratio Chart on Summary 3 ---
        ws_summary3 = writer.sheets['Summary 3']

        # Identify the number of data rows
        num_data_rows = len(weekly_summary_df)

        # Row/column bounds (1-based indexing for openpyxl)
        min_row = 2  # data starts after header
        max_row = min_row + num_data_rows - 1
        min_col_x = 1  # Column A (week)
        header_row = [cell.value for cell in ws_summary3[1]]
        if "Win Ratio (Win#/# of Trades)" in header_row:
            min_col_y = header_row.index("Win Ratio (Win#/# of Trades)") + 1

        # Create references
        cats = Reference(ws_summary3, min_col=min_col_x, max_col=min_col_x, min_row=min_row, max_row=max_row)
        data = Reference(ws_summary3, min_col=min_col_y, max_col=min_col_y, min_row=min_row, max_row=max_row)

        # Create and configure chart
        chart = LineChart()
        chart.title = "Weekly Win Ratio Over Time"
        chart.y_axis.title = "Win Ratio"
        chart.x_axis.title = "Week"
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # Optional: name the series
        chart.series[0].title = SeriesLabel(v="Weekly Win Ratio")

        # Insert chart into the sheet (e.g., cell G2)
        ws_summary3.add_chart(chart, "G2")

        # --- Chart generation (fully fixed) ---
        ws = writer.sheets['Summary 2']

        # Get valid month columns (exclude 'Total', 'Average', None)
        header_row = [cell.value for cell in ws[1]]
        valid_month_cols = []
        for col_idx, label in enumerate(header_row, start=1):
            if isinstance(label, str) and '-' in label and not label.lower().strip() in ['total', 'average']:
                valid_month_cols.append(col_idx)

        if not valid_month_cols:
            print("❌ No valid month columns found.")
            return

        min_col = min(valid_month_cols)
        max_col = max(valid_month_cols)

        # Helper to find row index by label in column A
        def find_row_index(label):
            for row in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=1).value
                if cell_val and str(cell_val).strip() == label:
                    return row
            return None

        # 1. Chart: cumulative_return_pct
        cumret_row_idx = find_row_index('cumulative_return_pct')
        if cumret_row_idx:
            data = Reference(ws, min_col=min_col, max_col=max_col,
                             min_row=cumret_row_idx, max_row=cumret_row_idx)
            cats = Reference(ws, min_col=min_col, max_col=max_col,
                             min_row=1, max_row=1)

            chart = LineChart()
            chart.title = "cumulative_return_pct"
            chart.style = 13
            chart.y_axis.title = "Cumulative %"
            chart.x_axis.title = "Month"
            chart.add_data(data, titles_from_data=False, from_rows=True)
            chart.set_categories(cats)
            chart.series[0].title = SeriesLabel(v="Cumulative Return (%)")

            ws.add_chart(chart, f"B{cumret_row_idx + 3}")
        else:
            print("❌ 'cumulative_return_pct' row not found.")

        # 2. Chart: Win Ratio
        winratio_row_idx = find_row_index('Win Ratio (Win/Total # Trade)')
        if winratio_row_idx:
            data = Reference(ws, min_col=min_col, max_col=max_col,
                             min_row=winratio_row_idx, max_row=winratio_row_idx)
            cats = Reference(ws, min_col=min_col, max_col=max_col,
                             min_row=1, max_row=1)

            chart2 = LineChart()
            chart2.title = "Win Ratio Over Time"
            chart2.style = 14
            chart2.y_axis.title = "Win Ratio"
            chart2.x_axis.title = "Month"
            chart2.add_data(data, titles_from_data=False, from_rows=True)
            chart2.set_categories(cats)
            chart2.series[0].title = SeriesLabel(v="Win Ratio")

            ws.add_chart(chart2, f"B{winratio_row_idx + 20}")
        else:
            print("❌ 'Win Ratio (Win/Total # Trade)' row not found.")

        # Calculate global stats
        max_win, max_loss = calculate_max_streaks(df_results)
        max_held_stocks = held_counts.max()

        # Append below the table
        worksheet = writer.sheets['Summary 2']
        start_row = monthly_summary_df.shape[0] + 3  # Leave 2 blank rows after the table

        worksheet.cell(row=start_row, column=1, value="Max consecutive days of winning")
        worksheet.cell(row=start_row, column=2, value=int(max_win))

        worksheet.cell(row=start_row + 1, column=1, value="Max consecutive days of losing")
        worksheet.cell(row=start_row + 1, column=2, value=int(max_loss))

        worksheet.cell(row=start_row + 2, column=1, value="Max number of stocks held at the same time")
        worksheet.cell(row=start_row + 2, column=2, value=int(max_held_stocks))

    print(f"✅ Backtest completed. Results saved to: {output_path}")


if __name__ == '__main__':
    main()